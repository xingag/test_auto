#!/usr/bin/env python  
# encoding: utf-8  

""" 
@version: v1.0 
@author: xag 
@license: Apache Licence  
@contact: xinganguo@gmail.com 
@site: https://github.com/xingag 
@software: PyCharm 
@file: demo.py 
@time: 2020/10/21 11:55 
@description：使用openpyxl处理Excel文件
"""

# 依赖
# pip3 install openpyxl

# 优点：支持xlsx
# 缺点：不支持xlx

import openpyxl
from utils.openpyxl_util import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font,Alignment


class ExcelF(object):
    def __init__(self):
        self.file_path = "./output.xlsx"

    def start(self):
        # 1、读取本地Excel文件
        # self.read_from_excel(self.file_path)

        # 2、保存数据到Excel文件中
        # self.write_data_to_excel()

        # 3.1、修改单元格样式
        # self.modify_cell_style1()

        # 3.3、修改本地Excel文件中的数据
        # self.modify_excel(self.file_path)

        # 4、openpyxl进阶使用
        self.advance_use(self.file_path)

    def write_data_to_excel(self):
        """
        保存数据到Excel文件中
        :return:
        """
        # 创建一个Excel工作簿
        # 注意：每次新建一个Excel文件，都会默认生成一个名称为【Sheet】的工作表Sheet
        wb = openpyxl.Workbook()

        # 查看文件中所有的sheet
        sheets = get_all_sheet(wb)
        # print(sheets)

        # 创建一个新的sheet，默认被插到到尾部
        # new_sheet = wb.create_sheet('新的Sheet')
        # 也可以通过第二个参数：index来指定插入的位置
        # 比如：插入到开头
        new_sheet = wb.create_sheet('新的Sheet', 0)

        # 设置Sheet的背景色(红色)
        set_sheet_bg_color(new_sheet, 'FF0000')
        # 复制一个Sheet
        # copy_a_sheet(wb, new_sheet)

        # 选中某一个sheet
        # sheet = get_sheet_by_name(wb, "Sheet")
        # print(sheet)

        # 向单元格中写入数据
        # 两种方式，包含：行列索引、字母位置
        write_value_to_cell_with_num(new_sheet, 1, 1, "姓名")
        write_value_to_cell_with_index_str(new_sheet, "B1", "年龄")

        # 合并、分离单元格
        # 下面方式二选一
        # merge_or_unmerge_cells_with_str(new_sheet, 'A1', 'B2', True)
        # merge_or_unmerge_cells_with_str(new_sheet, 'A1', 'B2', False)

        # merge_or_unmerge_cells_with_num(new_sheet, 1, 1, 2, 2, True)
        # merge_or_unmerge_cells_with_num(new_sheet, 1, 1, 2, 2, False)

        # 插入图片
        # 下面方式二选一
        # insert_img_to_cell_with_str(new_sheet, './1.png', 'A1')
        insert_img_to_cell_with_num(new_sheet, './1.png', 1, 1)

        # 注意：必须要写入，才能真实的保存到文件中
        wb.template = False
        wb.save('new.xlsx')

    def read_from_excel(self, file_path):
        """
        读取本地Excel文件
        :return:
        """
        # 加载本地的Excel文件
        wb = openpyxl.load_workbook(file_path)

        # =======================Sheet======================================
        # 查看所有的sheet名称
        sheet_names = get_all_sheet_names(wb)
        print("sheet包含：", sheet_names)

        # 所有的sheet
        sheets = get_all_sheet(wb)
        print("所有的sheet：", sheets)

        # 当前选择的sheet
        current_sheet = get_current_sheet(wb)
        print("current sheet:", current_sheet)

        # 通过sheet名称去查找某一个sheet
        one_sheet = get_sheet_by_name(wb, "第一个Sheet")
        print('查询的sheet为:', one_sheet)

        # =======================Cell单元格======================================
        # 选择某一个sheet
        sheet = get_sheet_by_name(wb, "第一个Sheet")

        # 行数和列数
        row_count, column_count = get_row_and_column_num(sheet)
        print('行数和列数分别为：', row_count, column_count)
        # 行数据
        print('行数据：', get_row_and_column_cells(sheet)[0])
        # 列数据
        print('列数据：', get_row_and_column_cells(sheet)[1])

        # 获取某一个sheet下，所有数据
        for row_index in range(row_count):
            for column_index in range(column_count):
                # 单元格对象
                cell = get_cell(sheet, row_index + 1, column_index + 1)
                # 获取单元格的数据及数据类型
                cell_value, cell_type = get_cell_value_and_type(cell)
                print('row_index:', row_index, ',column_index:', column_index, '单元格的数据值为:', cell_value, "，数据类型为：",
                      cell_type)

        # ============================行和列===================================
        # 获取某一行数据(元祖)
        # 比如：获取第一行数据
        rows = get_row_cells_by_index(sheet, 1)
        print('第一行的数据：', rows)

        # 获取某一列的数据
        # 比如：获取第二列数据
        columns = get_column_cells_by_index(sheet, 2)
        print("第二列的数据为：", columns)

        # 行范围的值
        # 比如：从第二行到第三行
        rows_range = get_rows_by_range(sheet, 2, 3)
        print(rows_range)

        # 列范围的值
        # 比如：获取第二列到第三列
        columns_range = get_columns_by_range(sheet, 2, 3)
        print(columns_range)

    def modify_cell_style1(self):
        """
        修改单元格样式
        openpyxl提供的格式控制方法可以实现对单元格属性所有基本操作
        :return:
        """
        # 字体格式
        # 指定字体类型、大小、是否加粗、颜色等
        font0 = Font(name='Calibri',
                     size=20,
                     bold=False,
                     italic=False,
                     vertAlign=None,  # Maybe：'baseline', 'superscript', 'subscript'
                     underline='none',  # Maybe：'single','double','singleAccounting','doubleAccounting'
                     strike=False,
                     color='FF00FF00')

        # 单元格填充
        fill0 = PatternFill(fill_type=None,
                            # Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
                            start_color='FFFFFFFF',
                            end_color='FF000000')

        # 边框
        border0 = Border(left=Side(border_style=None, color='FF000000'),
                         # style Maybe：'mediumDashDotDot', 'dotted', 'thick', 'medium', 'dashDotDot', 'double', 'dashed', 'mediumDashed', 'dashDot', 'mediumDashDot', 'hair', 'slantDashDot', 'thin'
                         right=Side(border_style=None, color='FF000000'),
                         top=Side(border_style=None, color='FF000000'),
                         bottom=Side(border_style=None, color='FF000000'),
                         diagonal=Side(border_style=None, color='FF000000'),
                         diagonal_direction=0,
                         outline=Side(border_style=None, color='FF000000'),
                         vertical=Side(border_style=None, color='FF000000'),
                         horizontal=Side(border_style=None, color='FF000000')
                         )

        # 单元格对齐方式
        alignment0 = Alignment(horizontal='center',
                               # Maybe:'centerContinuous', 'fill', 'right', 'distributed', 'justify', 'general', 'center', 'left'
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)

        # 表格保护
        protection0 = Protection(locked=True,
                                 hidden=False)

        # 创建格式
        style0 = NamedStyle(name='style_example')

        # 格式赋值
        style0.font = font0
        style0.alignment = alignment0
        style0.border = border0
        style0.fill = fill0
        style0.Protection = protection0

        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        write_value_to_cell_with_num(sheet, 1, 1, 'xag')

        # 设置属性样式
        sheet['A1'].font = font0
        sheet['A1'].alignment = alignment0
        sheet.cell(row=1, column=1).border = border0

        # 按名称调用
        # sheet['A1'].style = style0
        # sheet['A1'].style = 'style_example'
        # sheet.cell(row=1, column=1).style = style0

        wb.template = False
        wb.save('new.xlsx')

    def modify_excel(self, file_path):
        """
        修改本地Excel文件中数据
        :param file_path:
        :return:
        """
        # 读取本地Excel文件
        wb = openpyxl.load_workbook(file_path)

        # 读取某一个sheet
        sheet = wb['第一个Sheet']
        print(sheet)

        # 直接修改某一个单元格的数据
        write_value_to_cell_with_num(sheet, 1, 1, '姓名1')

        # 保存并覆盖
        wb.save(file_path)

    def advance_use(self, file_path):
        """
        openpyxl进阶使用
        :return:
        """
        # 1、隐藏的Sheet
        # 打开本地的Excel文件
        wb = openpyxl.load_workbook(file_path)

        # 所有可见的sheet
        sheets_visiable = get_all_visiable_sheets(wb)
        # 所有隐藏的sheet
        sheets_hidden = get_all_hidden_sheets(wb)

        # print('所有显示的sheet:', sheets_visiable)
        # print('所有隐藏的sheet:', sheets_hidden)

        # 2、查看某一个sheet中可见的行、列及隐藏的行列
        current_sheet = get_sheet_by_name(wb, '第一个Sheet')
        # print(current_sheet)

        # 获取某一行数据
        row = get_row(current_sheet, 0)
        # print('row:', row)

        # 获取某一列数据
        column = get_column(current_sheet, 0)
        # print('column:', column)

        # 2.1 获取隐藏的行索引列表
        # all_rows_hidden = get_all_rows_index(current_sheet, True)
        # print('所有隐藏的行列表为：', all_rows_hidden)
        # all_rows_visiable = get_all_rows_index(current_sheet, False)
        # print('所有显示的行列表为：', all_rows_visiable)
        #
        # all_columns_hidden = get_all_columns_index(current_sheet, True)
        # print('所有隐藏的行列表为：', all_columns_hidden)
        # all_columns_visiable = get_all_columns_index(current_sheet, False)
        # print('所有显示的行列表为：', all_columns_visiable)

        # 2.2 控制行、列的显示隐藏
        # 注意：设置完行列显示、隐藏后，需要保存后才会生效
        # set_rows_visiable(current_sheet, 1, 1)
        # set_rows_visiable(current_sheet, 3, 4)
        # set_rows_hidden(current_sheet, 1, 1)
        # set_rows_hidden(current_sheet, 3, 4)

        # set_columns_hidden_by(current_sheet,1,1, False)
        # set_columns_visiable(current_sheet, 1, 1, False)
        # set_columns_visiable(current_sheet, 'C', 'C', True)
        # set_columns_hidden(current_sheet, 'C', 'C', True)

        # 4、获取字体样式，包含：颜色、大小等
        # 获取单元格字体颜色
        cell_font_color = get_cell_font_color(current_sheet, 1, 1)
        print('字体颜色：', cell_font_color)
        # 获取单元格背景颜色
        cell_bg_color = get_cell_bg_color(current_sheet, 1, 1)
        print('背景颜色:', cell_bg_color)

        # 获取字体属性，包含：字体名称、大小、是否粗体
        font_properties = get_cell_font_properties(current_sheet, 1, 1)
        # 字体名称
        font_name = font_properties.name
        # 字体大小
        font_size = font_properties.size
        # 是否为粗体，True或者False
        font_bold = font_properties.bold
        # 是否为斜体
        font_italic = font_properties.italic
        print("字体名称：", font_name, ",字体大小：", font_size, ",是否粗体：", font_bold, ",是否斜体:", font_italic)

        # 保存文件
        wb.save(file_path)


if __name__ == '__main__':
    excelF = ExcelF()
    excelF.start()
