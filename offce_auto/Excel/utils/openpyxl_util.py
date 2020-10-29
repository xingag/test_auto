#!/usr/bin/env python  
# encoding: utf-8  

""" 
@version: v1.0 
@author: xag 
@license: Apache Licence  
@contact: xinganguo@gmail.com 
@site: https://github.com/xingag 
@software: PyCharm 
@file: openpyxl_util.py 
@time: 2020/10/21 12:03 
@description：openpyxl工具类
"""
import time
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image


def get_all_sheet_names(wb):
    """
    获取所有sheet的名称
    :param wb:
    :return:
    """
    # sheet名称列表
    sheet_names = wb.sheetnames
    return sheet_names


def get_all_sheet(wb):
    """
    获取所有的sheet
    :param wb:
    :return:
    """
    # sheet名称列表
    sheet_names = get_all_sheet_names(wb)

    # 所有sheet
    sheets = []
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        sheets.append(sheet)

    return sheets


def get_current_sheet(wb):
    """
    获取当前选择的sheet，默认是最后一个sheet
    :param wb:
    :return:
    """
    # 当前选中的sheet
    current_sheet = wb.active

    return current_sheet


def get_sheet_by_name(wb, sheet_name):
    """
    通过sheetname去查找某一个sheet
    :param wb:
    :param sheet_name:
    :return:
    """
    sheet_names = get_all_sheet_names(wb)
    if sheet_name in sheet_names:
        result = wb[sheet_name]
    else:
        result = None
    return result


def get_row_and_column_num(sheet):
    """
    获取sheet的行数和列数
    :param sheet:
    :return:
    """
    # 行数
    row_count = sheet.max_row
    # 列数
    column_count = sheet.max_column

    return row_count, column_count


def get_row_and_column_cells(sheet):
    """
    从sheet中获取所有的行数据、列数据
    :param sheet:
    :return:
    """
    # 行数据
    row_cells = sheet.rows

    # 列数据
    column_cells = sheet.columns

    return row_cells, column_cells


def get_row(sheet, row_index):
    """
    获取某一行
    :param sheet:
    :param row_index:行索引
    :return:
    """
    rows = list(get_row_and_column_cells(sheet)[0])
    return rows[row_index]


def get_all_rows_index(sheet, hidden_or_visiable):
    """
    获取所有隐藏/显示的行
    :param hidden_or_visiable:  True:隐藏；False：显示
    :param sheet:
    :return:
    """
    # 遍历行
    # 隐藏的索引
    hidden_indexs = []

    # 所有隐藏的行索引
    for row_index, rowDimension in sheet.row_dimensions.items():
        if rowDimension.hidden:
            hidden_indexs.append(row_index)

    # 所有显示的行索引
    visiable_indexs = [index + 1 for index in range(get_row_and_column_num(sheet)[0]) if index + 1 not in hidden_indexs]

    # 隐藏或者显示的行索引列表
    return hidden_indexs if hidden_or_visiable else visiable_indexs


def get_all_columns_index(sheet, hidden_or_visiable):
    """
    获取所有隐藏/显示的列
    :param sheet:
    :param hidden_or_visiable=True True:隐藏；False：显示
    :return:
    """
    # 遍历列
    # 隐藏的索引
    hidden_indexs = []

    # 所有隐藏的行索引
    for column_index, columnDimension in sheet.column_dimensions.items():
        if columnDimension.hidden:
            hidden_indexs.append(column_index)

    # 所有显示的行索引
    visiable_indexs = [index + 1 for index in range(get_row_and_column_num(sheet)[1]) if index + 1 not in hidden_indexs]

    # 隐藏或者显示的行索引列表
    return hidden_indexs if hidden_or_visiable else visiable_indexs


def get_column(sheet, column_index):
    """
    获取某一列
    :param sheet:
    :param column_index: 列索引
    :return:
    """
    columns = list(get_row_and_column_cells(sheet)[1])
    return columns[column_index]


def get_cell(sheet, row_index, column_index):
    """
    获取单元格
    :param sheet:
    :param row_index:
    :param column_index:
    :return:
    """
    # openpyxl索引都是从1开始计数，这与xlrd有所不同
    # 获取某一个单元格（二选一）
    # 比如：获取A1单元格的数据，即第一个行、第一列的数据
    # cell_one = sheet['A1']
    cell_one = sheet.cell(row=row_index, column=column_index)
    return cell_one


def get_cell_value_and_type(cell):
    """
    获取某一个cell的内容及数据类型
    :param cell:
    :return:
    """
    # 单元格的值
    cell_value = cell.value
    # 单元格的类型
    cell_type = get_cell_value_type(cell_value)

    return cell_value, cell_type


def get_cell_value_type(cell_value):
    """
    获取数据类型
    :param cell_value:
    :return:
    """
    # 其中
    # 0：空
    # 1：数字
    # 2：字符串
    # 3：日期
    # 4：其他
    if not cell_value:
        cell_type = 0
    elif isinstance(cell_value, int) or isinstance(cell_value, float):
        cell_type = 1
    elif isinstance(cell_value, str):
        cell_type = 2
    elif isinstance(cell_value, datetime.datetime):
        cell_type = 3
    else:
        cell_type = 4
    return cell_type


def column_num_to_str(num):
    """
    Excel索引列从数字转为字母
    :param num:
    :return:
    """
    return get_column_letter(num)


def column_str_to_num(str):
    """
    Excel索引列，从字母转为数字
    :param str:
    :return:
    """
    return column_index_from_string(str)


def get_row_cells_by_index(sheet, row_index):
    """
    通过行索引，获取某一行的单元格
    :param row_index:
    :return:
    """
    # 注意：第一列从1开始
    row_cells = sheet[row_index]
    return row_cells


def get_column_cells_by_index(sheet, column_index):
    """
    通过列索引，获取某一列的单元格
    """
    # 数字转为字母
    column_index_str = column_num_to_str(column_index)
    # 获取某一列的数据
    column_cells = sheet[column_index_str]
    return column_cells


def get_rows_by_range(sheet, row_index_start, row_index_end):
    """
    通过范围去选择行范围
    比如：选择第2行到第4行的所有数据，返回值为元组
    :param sheet:
    :param row_index_start:
    :param row_index_end:
    :return:
    """
    rows_range = sheet[row_index_start:row_index_end]
    return rows_range


def get_columns_by_range(sheet, column_index_start, column_index_end):
    """
    通过范围去选择列范围
    比如：选择第2列到第4列的所有数据，返回值为元组
    :param sheet:
    :param column_index_start:
    :param column_index_end:
    :return:
    """
    columns_range = sheet[column_num_to_str(column_index_start):column_num_to_str(column_index_end)]
    return columns_range


def set_sheet_bg_color(sheet, rgb_value):
    """
    设置Sheet标签的颜色
    :param rgb_value:
    :return:
    """
    # 设置Sheet底部按钮的颜色（RRGGBB）
    sheet.sheet_properties.tabColor = rgb_value


def copy_a_sheet(wb, sheet):
    """
    复制一个Sheet
    :param wb:
    :param sheet:
    :return:
    """
    # 复制一个工作表Sheet
    copy_sheet = wb.copy_worksheet(sheet)
    return copy_sheet


def write_value_to_cell_with_num(sheet, row_index, column_index, value):
    """
    按行索引、列索引写入数据
    :param shell:
    :param row_index: 行索引
    :param column_index: 列索引
    :param value:
    :return:
    """
    # 二选一
    sheet.cell(row=row_index, column=column_index, value=value)
    # shell.cell(row=row_index, column=column_index).value = value


def write_value_to_cell_with_index_str(sheet, index_str, value):
    """
    按字母位置，写入数据到对应单元格
    :param shell:
    :param index_str: 字母对应的单元格位置
    :param value:
    :return:
    """
    sheet[index_str] = value
    # sheet['A2'] = 4


def merge_or_unmerge_cells_with_str(sheet, start_index_str, end_index_str, merge_or_unmerge):
    """
     合并单元格
    :param shell:
    :param start_index_str: 开始位置
    :param end_index_str: 结束位置
    :param merge_or_unmerge:合并还是分离，true：合并；false：分离
    :return:
    """
    # 合并单元格
    # 比如：合并第1行1列 - 第2行2列
    # sheet.merge_cells('A1:B2')
    if merge_or_unmerge:
        sheet.merge_cells('%s:%s' % (start_index_str, end_index_str))
    else:
        sheet.unmerge_cells('%s:%s' % (start_index_str, end_index_str))


def merge_or_unmerge_cells_with_num(sheet, start_row_index, start_column_index, end_row_index, end_column_index,
                                    merge_or_unmerge):
    """
     合并单元格
    :param sheet:
    :param start_row_index:
    :param start_column_index:
    :param end_row_index:
    :param end_column_index:
    :param merge_or_unmerge:合并还是分离，true：合并；false：分离
    :return:
    """
    if merge_or_unmerge:
        sheet.merge_cells(start_row=start_row_index, start_column=start_column_index, end_row=end_row_index,
                          end_column=end_column_index)
    else:
        sheet.unmerge_cells(start_row=start_row_index, start_column=start_column_index, end_row=end_row_index,
                            end_column=end_column_index)


def insert_img_to_cell_with_str(sheet, image_path, index_str):
    """
    往单元格中插入图片
    :param sheet:
    :param image_path:
    :param index_str:
    :return:
    """
    sheet.add_image((image_path), index_str)


def insert_img_to_cell_with_num(sheet, image_path, row_index, column_index):
    """
    往单元格中插入图片
    :param sheet:
    :param image_path:
    :param row_index:
    :param column_index:
    :return:
    """
    # 通过行索引、列索引，获取到字母索引
    index_str = column_num_to_str(column_index) + str(row_index)
    insert_img_to_cell_with_str(sheet, image_path, index_str)


def get_all_visiable_sheets(wb):
    """
    获取工作簿中所有可见的sheet
    :param wb:
    :return:
    """
    return [sheet for sheet in get_all_sheet(wb) if sheet.sheet_state == 'visible']


def get_all_hidden_sheets(wb):
    """
    获取工作簿中所有隐藏的sheet
    :param wb:
    :return:
    """
    return [sheet for sheet in get_all_sheet(wb) if sheet.sheet_state == 'hidden']

def set_rows_visiable(sheet, start_row_index, end_row_index):
    """
    设置行显示
    :param sheet:
    :param start_row_index: 开始位置（数字）
    :param end_row_index: 结束位置（数字）
    :return:
    """
    sheet.row_dimensions.group(start_row_index, end_row_index, hidden=False)


def set_rows_hidden(sheet, start_row_index, end_row_index):
    """
    设置行隐藏
    :param sheet:
    :param start_row_index:
    :param end_row_index:
    :return:
    """
    sheet.row_dimensions.group(start_row_index, end_row_index, hidden=True)


def set_columns_visiable(sheet, start_column_index, end_column_index, str_or_num):
    """
    设置列显示
    :param sheet:
    :param start_column_index: 开始位置（字母）
    :param end_column_index: 结束位置（字母）
    :param str_or_num 字母索引还是数字索引
    :return:
    """
    # 如果是数字，转为字母索引
    if not str_or_num:
        # 列索引从数字转为字母
        start_column_index = column_num_to_str(start_column_index)
        end_column_index = column_num_to_str(end_column_index)
    sheet.column_dimensions.group(start_column_index, end_column_index, hidden=False)


def set_columns_hidden(sheet, start_column_index, end_column_index, str_or_num):
    """
    设置行隐藏
    :param sheet:
    :param start_column_index:开始位置（字母）
    :param end_column_index:结束位置（字母）
    :param str_or_num True:字母；False：数字
    :return:
    """
    # 如果是数字，转为字母索引
    if not str_or_num:
        # 列索引从数字转为字母
        start_column_index = column_num_to_str(start_column_index)
        end_column_index = column_num_to_str(end_column_index)
    sheet.column_dimensions.group(start_column_index, end_column_index, hidden=True)


def get_cell_font_color(sheet, row_index, column_index):
    """
    获取单元格字体的颜色
    :param sheet:
    :param row_index:行索引
    :param column_index:列索引
    :return:
    """
    cell_color = sheet.cell(row_index, column_index).font.color
    if cell_color:
        return sheet.cell(row_index, column_index).font.color.rgb
    else:
        # 颜色不存在，可能单元格没有数据
        return None


def get_cell_bg_color(sheet, row_index, column_index):
    """
    获取单元格背景的颜色
    :param sheet:
    :param row_index:行索引
    :param column_index:列索引
    :return:
    """
    return sheet.cell(row_index, column_index).fill.fgColor.rgb


def get_cell_font_properties(sheet, row_index, column_index):
    """
    获取单元格字体的字体属性，包含：名称、字体大小、是否加粗、是否斜体等
    :param sheet:
    :param row_index:
    :param column_index:
    :return:
    """
    return sheet.cell(row_index, column_index).font
